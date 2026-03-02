"""
Genera la plantilla Excel para importar servicios.
"""
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from supabase import create_client

SUPABASE_URL = 'https://drjnffoyzuploatfcltf.supabase.co'
SUPABASE_KEY = 'eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6ImRyam5mZm95enVwbG9hdGZjbHRmIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3MDkzMTA5MywiZXhwIjoyMDg2NTA3MDkzfQ.dJbTqUCQh6T4H2r6ThBm-CcuzSw0diiGNKsdIDFoyiY'

sb  = create_client(SUPABASE_URL, SUPABASE_KEY)
emp = sb.table('empleados').select('id, nombre, apellidos').eq('activo', True).order('apellidos').execute().data
veh = sb.table('vehiculos').select('id, matricula, marca, modelo').execute().data

# Listas para validación desplegable
lista_emp = [f"{e['apellidos']}, {e['nombre']}" for e in emp]
lista_veh = [f"{v['matricula']} - {v['marca']} {v['modelo']}" for v in veh]

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Servicios"

# ── Estilos ────────────────────────────────────────────────────
CABECERA_COLS = {
    # (col_excel, ancho, titulo, descripcion, grupo)
    # IDENTIFICACIÓN DEL SERVICIO
    "A":  (22, "CODIGO_SERVICIO",        "Código único del servicio (ej: SRV-021)",              "SERVICIO"),
    "B":  (35, "DESCRIPCION",            "Descripción del servicio",                              "SERVICIO"),
    "C":  (28, "ZONA_LOCALIDAD",         "Zona o localidad del reparto",                          "SERVICIO"),
    "D":  (18, "TIPO_SERVICIO",          "Tipo: reparto / recogida / mixto",                      "SERVICIO"),
    "E":  (18, "FECHA_INICIO_CONTRATO",  "Fecha inicio del contrato (DD/MM/AAAA)",                "SERVICIO"),
    "F":  (18, "FECHA_FIN_CONTRATO",     "Fecha fin contrato si aplica (DD/MM/AAAA)",             "SERVICIO"),
    "G":  (18, "DIAS_SERVICIO",          "Ej: Lunes-Viernes / Lunes-Sábado / Diario",             "SERVICIO"),
    "H":  (18, "HORARIO_INICIO",         "Hora inicio (ej: 08:00)",                               "SERVICIO"),
    "I":  (18, "HORARIO_FIN",            "Hora fin (ej: 14:00)",                                  "SERVICIO"),
    # RECURSOS ASIGNADOS
    "J":  (40, "EMPLEADO_BASE",          "Apellidos, Nombre del empleado asignado",               "RECURSOS"),
    "K":  (32, "VEHICULO_BASE",          "Matrícula - Marca Modelo del vehículo",                 "RECURSOS"),
    # EMPRESA CLIENTE
    "L":  (35, "EMPRESA_NOMBRE",         "Nombre o razón social de la empresa cliente",           "EMPRESA"),
    "M":  (18, "EMPRESA_CIF",            "CIF / NIF de la empresa cliente",                       "EMPRESA"),
    "N":  (40, "EMPRESA_DIRECCION",      "Dirección fiscal completa",                             "EMPRESA"),
    "O":  (12, "EMPRESA_CP",             "Código postal",                                         "EMPRESA"),
    "P":  (25, "EMPRESA_CIUDAD",         "Ciudad",                                                "EMPRESA"),
    "Q":  (20, "EMPRESA_PROVINCIA",      "Provincia",                                             "EMPRESA"),
    "R":  (20, "EMPRESA_PAIS",           "País (por defecto España)",                             "EMPRESA"),
    # CONTACTO PRINCIPAL
    "S":  (30, "CONTACTO_NOMBRE",        "Nombre y apellidos del contacto principal",             "CONTACTO"),
    "T":  (25, "CONTACTO_CARGO",         "Cargo o puesto del contacto",                          "CONTACTO"),
    "U":  (30, "CONTACTO_EMAIL",         "Email del contacto principal",                         "CONTACTO"),
    "V":  (18, "CONTACTO_TELEFONO",      "Teléfono del contacto principal",                      "CONTACTO"),
    "W":  (18, "CONTACTO_MOVIL",         "Móvil del contacto",                                   "CONTACTO"),
    # CONTACTO SECUNDARIO / URGENCIAS
    "X":  (30, "CONTACTO2_NOMBRE",       "Nombre contacto secundario / urgencias",               "CONTACTO2"),
    "Y":  (30, "CONTACTO2_EMAIL",        "Email contacto secundario",                            "CONTACTO2"),
    "Z":  (18, "CONTACTO2_TELEFONO",     "Teléfono contacto secundario",                         "CONTACTO2"),
    # FACTURACIÓN
    "AA": (35, "FACTURACION_EMAIL",      "Email para envío de facturas",                         "FACTURACION"),
    "AB": (20, "FACTURACION_FORMA_PAGO", "Forma de pago (ej: transferencia 30d)",                "FACTURACION"),
    "AC": (15, "TARIFA_MENSUAL",         "Tarifa mensual del servicio (€)",                      "FACTURACION"),
    # OBSERVACIONES
    "AD": (50, "OBSERVACIONES",          "Notas, condiciones especiales u observaciones",        "NOTAS"),
}

COLORES_GRUPO = {
    "SERVICIO":     ("1E3A5F", "DBEAFE"),  # azul oscuro / azul claro
    "RECURSOS":     ("1B5E20", "DCFCE7"),  # verde oscuro / verde claro
    "EMPRESA":      ("4A148C", "EDE9FE"),  # morado oscuro / morado claro
    "CONTACTO":     ("B45309", "FEF3C7"),  # naranja oscuro / naranja claro
    "CONTACTO2":    ("9D174D", "FCE7F3"),  # rosa oscuro / rosa claro
    "FACTURACION":  ("0F766E", "CCFBF1"),  # teal oscuro / teal claro
    "NOTAS":        ("374151", "F3F4F6"),  # gris oscuro / gris claro
}

thin = Side(style='thin', color='CCCCCC')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

# ── Fila 1: Grupos ────────────────────────────────────────────
grupos_rango = {}
for col, (ancho, campo, desc, grupo) in CABECERA_COLS.items():
    if grupo not in grupos_rango:
        grupos_rango[grupo] = [col, col]
    else:
        grupos_rango[grupo][1] = col

for grupo, (c_inicio, c_fin) in grupos_rango.items():
    cell = ws[f"{c_inicio}1"]
    col_dark, col_light = COLORES_GRUPO[grupo]
    cell.value       = grupo
    cell.font        = Font(bold=True, color="FFFFFF", size=10)
    cell.fill        = PatternFill("solid", fgColor=col_dark)
    cell.alignment   = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border      = border
    if c_inicio != c_fin:
        ws.merge_cells(f"{c_inicio}1:{c_fin}1")

ws.row_dimensions[1].height = 24

# ── Fila 2: Nombre del campo ──────────────────────────────────
for col, (ancho, campo, desc, grupo) in CABECERA_COLS.items():
    col_dark, col_light = COLORES_GRUPO[grupo]
    cell = ws[f"{col}2"]
    cell.value     = campo
    cell.font      = Font(bold=True, color="FFFFFF", size=9)
    cell.fill      = PatternFill("solid", fgColor=col_dark)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border    = border

ws.row_dimensions[2].height = 30

# ── Fila 3: Descripción / instrucción ─────────────────────────
for col, (ancho, campo, desc, grupo) in CABECERA_COLS.items():
    col_dark, col_light = COLORES_GRUPO[grupo]
    cell = ws[f"{col}3"]
    cell.value     = desc
    cell.font      = Font(italic=True, size=8, color="444444")
    cell.fill      = PatternFill("solid", fgColor=col_light)
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    cell.border    = border

ws.row_dimensions[3].height = 36

# ── Anchos de columna ─────────────────────────────────────────
for col, (ancho, campo, desc, grupo) in CABECERA_COLS.items():
    ws.column_dimensions[col].width = ancho

# ── Filas de datos (4 en adelante) ───────────────────────────
FILAS_DATOS = 100
for fila in range(4, 4 + FILAS_DATOS):
    for col, (ancho, campo, desc, grupo) in CABECERA_COLS.items():
        col_dark, col_light = COLORES_GRUPO[grupo]
        cell = ws[f"{col}{fila}"]
        cell.fill      = PatternFill("solid", fgColor="FFFFFF")
        cell.border    = border
        cell.alignment = Alignment(vertical="center", wrap_text=False)
        # Alternar fila ligeramente
        if fila % 2 == 0:
            cell.fill = PatternFill("solid", fgColor="F9FAFB")

# ── Validación: empleados (col J) ─────────────────────────────
# Hoja oculta con listas
ws_listas = wb.create_sheet("_listas")
ws_listas.sheet_state = 'hidden'

for i, nombre in enumerate(lista_emp, start=1):
    ws_listas[f"A{i}"] = nombre
for i, mat in enumerate(lista_veh, start=1):
    ws_listas[f"B{i}"] = mat

# Tipos de servicio
for i, t in enumerate(["reparto", "recogida", "mixto", "almacenaje"], start=1):
    ws_listas[f"C{i}"] = t

dv_emp = DataValidation(
    type="list",
    formula1=f"_listas!$A$1:$A${len(lista_emp)}",
    allow_blank=True,
    showDropDown=False,
)
dv_veh = DataValidation(
    type="list",
    formula1=f"_listas!$B$1:$B${len(lista_veh)}",
    allow_blank=True,
    showDropDown=False,
)
dv_tipo = DataValidation(
    type="list",
    formula1=f"_listas!$C$1:$C$4",
    allow_blank=True,
    showDropDown=False,
)

ws.add_data_validation(dv_emp)
ws.add_data_validation(dv_veh)
ws.add_data_validation(dv_tipo)

dv_emp.sqref  = f"J4:J{3 + FILAS_DATOS}"
dv_veh.sqref  = f"K4:K{3 + FILAS_DATOS}"
dv_tipo.sqref = f"D4:D{3 + FILAS_DATOS}"

# ── Inmovilizar cabecera ──────────────────────────────────────
ws.freeze_panes = "A4"

# ── Hoja de instrucciones ─────────────────────────────────────
ws_help = wb.create_sheet("INSTRUCCIONES")
instrucciones = [
    ("INSTRUCCIONES DE USO", True, "1E3A5F"),
    ("", False, None),
    ("1. Rellena una fila por cada servicio/ruta.", False, None),
    ("2. CODIGO_SERVICIO: único por servicio (ej: SRV-021, SRV-022...)", False, None),
    ("3. EMPLEADO_BASE: usa el desplegable (col J) para elegir el empleado.", False, None),
    ("4. VEHICULO_BASE: usa el desplegable (col K) para elegir el vehículo.", False, None),
    ("5. TIPO_SERVICIO: desplegable con: reparto / recogida / mixto / almacenaje", False, None),
    ("6. Fechas en formato DD/MM/AAAA", False, None),
    ("7. Los campos con * son OBLIGATORIOS: CODIGO_SERVICIO, DESCRIPCION, EMPLEADO_BASE, VEHICULO_BASE", False, None),
    ("8. El resto de campos son opcionales pero ayudan a tener toda la info centralizada.", False, None),
    ("", False, None),
    ("CAMPOS OBLIGATORIOS (mínimo para crear el servicio):", True, "1B5E20"),
    ("  - CODIGO_SERVICIO", False, None),
    ("  - DESCRIPCION", False, None),
    ("  - EMPLEADO_BASE (desplegable col J)", False, None),
    ("  - VEHICULO_BASE (desplegable col K)", False, None),
    ("", False, None),
    ("COLORES DE LAS COLUMNAS:", True, "4A148C"),
    ("  AZUL     = Datos básicos del servicio", False, None),
    ("  VERDE    = Empleado y vehículo asignados", False, None),
    ("  MORADO   = Datos de la empresa cliente", False, None),
    ("  NARANJA  = Contacto principal", False, None),
    ("  ROSA     = Contacto secundario / urgencias", False, None),
    ("  TEAL     = Facturación", False, None),
    ("  GRIS     = Observaciones", False, None),
]
for i, (txt, bold, color) in enumerate(instrucciones, start=1):
    cell = ws_help[f"A{i}"]
    cell.value = txt
    if bold:
        cell.font = Font(bold=True, size=11, color=color or "000000")
    else:
        cell.font = Font(size=10)
ws_help.column_dimensions["A"].width = 80

# ── Guardar ───────────────────────────────────────────────────
nombre_archivo = "PLANTILLA_SERVICIOS.xlsx"
wb.save(nombre_archivo)
print(f"Plantilla generada: {nombre_archivo}")
print(f"  - {len(lista_emp)} empleados en desplegable")
print(f"  - {len(lista_veh)} vehículos en desplegable")
print(f"  - {FILAS_DATOS} filas de datos disponibles")
